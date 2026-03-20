from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from app.services.parsers.document_parser import DocumentParser


class XlsxParser:
    def parse(self, file_path: str | Path) -> dict[str, Any]:
        path = Path(file_path)

        workbook = openpyxl.load_workbook(path, data_only=True)
        text_parts: list[str] = []
        tables: list[dict[str, Any]] = []
        sections: list[dict[str, Any]] = []

        try:
            for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
                worksheet = workbook[sheet_name]

                data = worksheet.values
                rows = list(data)

                normalized_rows: list[list[str]] = []
                sheet_text_lines: list[str] = []

                for row in rows:
                    if row is None:
                        continue

                    normalized_row = [
                        "" if cell is None else str(cell).strip()
                        for cell in row
                    ]

                    if any(normalized_row):
                        normalized_rows.append(normalized_row)
                        sheet_text_lines.append(" | ".join(normalized_row))

                if normalized_rows:
                    tables.append(
                        {
                            "index": len(tables) + 1,
                            "title": f"Worksheet: {sheet_name}",
                            "sheet": sheet_name,
                            "rows": normalized_rows,
                        }
                    )

                    sections.append(
                        {
                            "index": sheet_index,
                            "title": sheet_name,
                            "content": "\n".join(sheet_text_lines).strip(),
                        }
                    )

                    text_parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(sheet_text_lines))

            for sheet_name in workbook.sheetnames:
                try:
                    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
                    if not df.empty:
                        _ = df.fillna("").astype(str)
                except Exception:
                    continue

        finally:
            workbook.close()

        merged_text = "\n\n".join(text_parts).strip()
        normalized_text = DocumentParser.normalize_text(merged_text)

        if not sections:
            sections = DocumentParser.split_text_into_blocks(normalized_text)

        return {
            "text": normalized_text,
            "tables": tables,
            "sections": sections,
        }